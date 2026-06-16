"""
Daily Summary Generator — Pending Transactions
Reads all translog CSV files from the input folder and produces Daily_Summary.xlsx.

Usage:
  python generate_daily_summary.py <workspace_dir> [output_filename]

  workspace_dir:    path containing input/ and output/ subfolders
  output_filename:  optional, defaults to Daily_Summary.xlsx

File naming convention:
  Input CSVs map to date labels via a date_map.json file in the workspace root,
  OR by reading the Trans Time column from each CSV (format dd/mm/yyyy).

  date_map.json example:
    { "translog_9720.csv": "260101", "translog_9848.csv": "260602" }

  If no date_map.json exists, dates are read from the Trans Time column.
"""

import sys
import os
import json
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.utils import get_column_letter

# ── Business rules ────────────────────────────────────────────────────────────
EXCLUDE_SUB = {
    'LINK', 'QUERY_BALANCE_USER', 'REGISTER_QUERY_BALANCE',
    'UNLINK', 'VERIFY_OTP', 'UNREGISTER_QUERY_BALANCE'
}
DETAIL_SUBS = [
    'DOMESTIC_TRANSFER_FUND_2_ACCOUNT',
    'QR_PUSH_PAYMENT',
    'WITHDRAW_BY_TOKEN'
]
LOW_PRIORITY_RC = {
    'CALL_BANK_CONN_EXCEPTION (-9202)',
    'BANK_CASH_TIMEOUT (-5015)',
    'BANK_SYSTEM_ERROR (-5007)',
    'BANK_NOT_CONNECTED (-5001)'
}

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
SECTION_FILL = PatternFill("solid", start_color="2E75B6")
TOTAL_FILL   = PatternFill("solid", start_color="BDD7EE")
ALT_FILL     = PatternFill("solid", start_color="D9E1F2")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
DOWN_FILL    = PatternFill("solid", start_color="FFC7CE")
UP_FILL      = PatternFill("solid", start_color="C6EFCE")
WARN_FILL    = PatternFill("solid", start_color="FFEB9C")

H_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=10)
S_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=11)
N_FONT     = Font(name="Arial", size=9)
B_FONT     = Font(name="Arial", bold=True, size=9)
T_FONT     = Font(name="Arial", bold=True, size=9,  color="FFFFFF")
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F3864")
SEC_FONT   = Font(name="Arial", bold=True, size=11, color="FFFFFF")
BODY_FONT  = Font(name="Arial", size=10)
DOWN_FONT  = Font(name="Arial", bold=True, size=10, color="9C0006")
UP_FONT    = Font(name="Arial", bold=True, size=10, color="375623")
WARN_FONT  = Font(name="Arial", bold=True, size=10, color="7D6608")

thin  = Side(style='thin',   color="AAAAAA")
med   = Side(style='medium', color="2E75B6")
thin2 = Side(style='thin',   color="CCCCCC")
T_BDR = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
M_BDR = Border(left=med,   right=med,   top=med,   bottom=med)
B2    = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
L  = Alignment(horizontal='left',   vertical='center')
R  = Alignment(horizontal='right',  vertical='center')
LW = Alignment(horizontal='left',   vertical='top',    wrap_text=True)


# ── Helpers ───────────────────────────────────────────────────────────────────

def hc(ws, r, c, v):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = H_FONT; cl.fill = HEADER_FILL; cl.alignment = C; cl.border = T_BDR

def tc(ws, r, c, v):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = T_FONT; cl.fill = HEADER_FILL; cl.alignment = R; cl.border = T_BDR

def sc(ws, r, c, v, ncols):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + ncols - 1)
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = S_FONT; cl.fill = SECTION_FILL; cl.alignment = L; cl.border = M_BDR


def write_daily_table(ws, start_row, start_col, row_vals, dates, lookup, row_label, sort_date):
    """
    Layout: row_label | Grand Total | date1 | date2 | ...
    Rows sorted descending by sort_date column.
    """
    gt_col   = start_col + 1
    data_col = start_col + 2

    hc(ws, start_row, start_col, row_label)
    hc(ws, start_row, gt_col, "Grand Total")
    for j, d in enumerate(dates):
        hc(ws, start_row, data_col + j, d)

    sorted_rows = sorted(row_vals, key=lambda x: lookup.get((x, sort_date), 0), reverse=True)

    for i, rv in enumerate(sorted_rows):
        r = start_row + 1 + i
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        cl = ws.cell(row=r, column=start_col, value=str(rv))
        cl.font = B_FONT; cl.fill = fill; cl.alignment = L; cl.border = T_BDR
        refs = []
        for j, d in enumerate(dates):
            c = data_col + j
            val = lookup.get((rv, d), 0)
            cl = ws.cell(row=r, column=c, value=val)
            cl.font = N_FONT; cl.fill = fill; cl.alignment = R; cl.border = T_BDR
            refs.append(get_column_letter(c))
        cl = ws.cell(row=r, column=gt_col, value="=" + "+".join(f"{x}{r}" for x in refs))
        cl.font = B_FONT; cl.fill = TOTAL_FILL; cl.alignment = R; cl.border = T_BDR

    tr = start_row + 1 + len(sorted_rows)
    cl = ws.cell(row=tr, column=start_col, value="Grand Total")
    cl.font = T_FONT; cl.fill = HEADER_FILL; cl.alignment = L; cl.border = T_BDR
    gl = get_column_letter(gt_col)
    tc(ws, tr, gt_col, f"=SUM({gl}{start_row+1}:{gl}{tr-1})")
    for j in range(len(dates)):
        c = data_col + j; l = get_column_letter(c)
        tc(ws, tr, c, f"=SUM({l}{start_row+1}:{l}{tr-1})")

    ws.column_dimensions[get_column_letter(start_col)].width = max(
        ws.column_dimensions[get_column_letter(start_col)].width or 0, 30)
    ws.column_dimensions[get_column_letter(gt_col)].width = max(
        ws.column_dimensions[get_column_letter(gt_col)].width or 0, 14)
    for j in range(len(dates)):
        c = data_col + j
        ws.column_dimensions[get_column_letter(c)].width = max(
            ws.column_dimensions[get_column_letter(c)].width or 0, 12)

    return tr + 2


# ── Tab builders ──────────────────────────────────────────────────────────────

def build_bcc_tab(ws, data, dates, sort_date):
    ws["A1"] = "Daily Summary – By Bank Connector Code"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1F3864"); ws["A1"].alignment = L

    bcc_all = sorted(data['Bank Connector Code'].unique())
    lk = data.groupby(['Bank Connector Code', '_date']).size().to_dict()
    current_row = write_daily_table(ws, 2, 1, bcc_all, dates, lk, "Bank Connector Code", sort_date)

    for st in sorted(data['Sub Trans Type'].unique()):
        d_st = data[data['Sub Trans Type'] == st]
        bcc_list = sorted(d_st['Bank Connector Code'].unique())
        rc_list  = sorted(d_st['ReturnCode (BC)'].unique())
        sc(ws, current_row, 1, f"▸  {st}", len(dates) + 3); current_row += 1

        ws.cell(row=current_row, column=1).value = "Bank Connector Code – Count Daily"
        ws.cell(row=current_row, column=1).font = Font(name="Arial", bold=True, size=10, color="2E75B6")
        ws.cell(row=current_row, column=1).alignment = L; current_row += 1
        lk_bcc = d_st.groupby(['Bank Connector Code', '_date']).size().to_dict()
        current_row = write_daily_table(ws, current_row, 1, bcc_list, dates, lk_bcc, "Bank Connector Code", sort_date)

        ws.cell(row=current_row, column=1).value = "ReturnCode (BC) – Count Daily"
        ws.cell(row=current_row, column=1).font = Font(name="Arial", bold=True, size=10, color="2E75B6")
        ws.cell(row=current_row, column=1).alignment = L; current_row += 1
        lk_rc = d_st.groupby(['ReturnCode (BC)', '_date']).size().to_dict()
        current_row = write_daily_table(ws, current_row, 1, rc_list, dates, lk_rc, "ReturnCode (BC)", sort_date)

    ws.freeze_panes = "C3"


def build_rc_tab(ws, data, dates, sort_date):
    ws["A1"] = "Daily Summary – ReturnCode (BC)"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1F3864"); ws["A1"].alignment = L

    rc_all = sorted(data['ReturnCode (BC)'].unique())
    lk = data.groupby(['ReturnCode (BC)', '_date']).size().to_dict()
    write_daily_table(ws, 2, 1, rc_all, dates, lk, "ReturnCode (BC)", sort_date)
    ws.freeze_panes = "C3"


def build_detail_tab(ws, data, dates, sort_date):
    d_det = data[data['Sub Trans Type'].isin(DETAIL_SUBS)]
    all_sub_types = sorted(data['Sub Trans Type'].unique())

    ws["A1"] = "Daily Summary – Sub Trans Type (Detail)"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1F3864"); ws["A1"].alignment = L
    ws.cell(row=2, column=1).value = "Overview – Sub Trans Type Count Daily"
    ws.cell(row=2, column=1).font = Font(name="Arial", bold=True, size=11, color="1F3864")
    ws.cell(row=2, column=1).alignment = L

    lk_overview = data.groupby(['Sub Trans Type', '_date']).size().to_dict()
    current_row = write_daily_table(ws, 3, 1, all_sub_types, dates, lk_overview, "Sub Trans Type", sort_date)

    for st in DETAIL_SUBS:
        if st not in d_det['Sub Trans Type'].values:
            continue
        d_st = d_det[d_det['Sub Trans Type'] == st]
        bcc_list = sorted(d_st['Bank Connector Code'].unique())
        rc_list  = sorted(d_st['ReturnCode (BC)'].unique())
        sc(ws, current_row, 1, f"▸  {st}", len(dates) + 3); current_row += 1

        ws.cell(row=current_row, column=1).value = "Bank Connector Code – Count Daily"
        ws.cell(row=current_row, column=1).font = Font(name="Arial", bold=True, size=10, color="2E75B6")
        ws.cell(row=current_row, column=1).alignment = L; current_row += 1
        lk_bcc = d_st.groupby(['Bank Connector Code', '_date']).size().to_dict()
        current_row = write_daily_table(ws, current_row, 1, bcc_list, dates, lk_bcc, "Bank Connector Code", sort_date)

        ws.cell(row=current_row, column=1).value = "ReturnCode (BC) – Count Daily"
        ws.cell(row=current_row, column=1).font = Font(name="Arial", bold=True, size=10, color="2E75B6")
        ws.cell(row=current_row, column=1).alignment = L; current_row += 1
        lk_rc = d_st.groupby(['ReturnCode (BC)', '_date']).size().to_dict()
        current_row = write_daily_table(ws, current_row, 1, rc_list, dates, lk_rc, "ReturnCode (BC)", sort_date)

    ws.freeze_panes = "C4"


def build_charts_tab(ws, data, dates, sort_date):
    def get_lookup(pivot, row_key):
        rows = sorted(pivot[row_key].unique())
        lk = {rv: {d: int(pivot[(pivot[row_key]==rv)&(pivot['_date']==d)]['n'].sum()) for d in dates} for rv in rows}
        rows = sorted(rows, key=lambda x: sum(lk[x].values()), reverse=True)
        return rows, lk

    bcc_pivot = data.groupby(['Bank Connector Code', '_date']).size().reset_index(name='n')
    rc_pivot  = data.groupby(['ReturnCode (BC)',      '_date']).size().reset_index(name='n')
    bcc_rows, bcc_lk = get_lookup(bcc_pivot, 'Bank Connector Code')
    rc_rows,  rc_lk  = get_lookup(rc_pivot,  'ReturnCode (BC)')

    def write_chart_data(ws, start_row, start_col, rows, dates, lk, title):
        ws.cell(row=start_row, column=start_col, value=title).font = Font(name="Arial", bold=True, size=11, color="1F3864")
        ws.cell(row=start_row, column=start_col).alignment = L
        ws.cell(row=start_row+1, column=start_col).fill = HEADER_FILL
        ws.cell(row=start_row+1, column=start_col).border = T_BDR
        for j, d in enumerate(dates):
            cl = ws.cell(row=start_row+1, column=start_col+1+j, value=d)
            cl.font = H_FONT; cl.fill = HEADER_FILL; cl.alignment = C; cl.border = T_BDR
        for i, rv in enumerate(rows):
            r = start_row+2+i
            ws.cell(row=r, column=start_col, value=str(rv)).border = T_BDR
            ws.cell(row=r, column=start_col).font = N_FONT
            for j, d in enumerate(dates):
                cl = ws.cell(row=r, column=start_col+1+j, value=lk[rv][d])
                cl.border = T_BDR; cl.font = N_FONT; cl.alignment = R
        ws.column_dimensions[get_column_letter(start_col)].width = 36
        for j in range(len(dates)):
            ws.column_dimensions[get_column_letter(start_col+1+j)].width = 12
        return start_row+2+len(rows)

    bcc_start = 2
    bcc_end = write_chart_data(ws, bcc_start, 1, bcc_rows, dates, bcc_lk, "Daily Pending – By Bank Connector Code")
    rc_start = bcc_end + 3
    rc_end = write_chart_data(ws, rc_start, 1, rc_rows, dates, rc_lk, "Daily Pending – By ReturnCode (BC)")

    def add_chart(ws, title, rows, data_start, anchor):
        chart = BarChart()
        chart.type = "col"; chart.grouping = "stacked"; chart.overlap = 100
        chart.title = title; chart.y_axis.title = "Count"; chart.x_axis.title = "Date"
        chart.style = 10; chart.width = 30; chart.height = 20
        cats = Reference(ws, min_col=2, max_col=1+len(dates), min_row=data_start+1, max_row=data_start+1)
        for i, rv in enumerate(rows):
            values = Reference(ws, min_col=2, max_col=1+len(dates), min_row=data_start+2+i, max_row=data_start+2+i)
            chart.series.append(Series(values, title=str(rv)))
        chart.set_categories(cats)
        ws.add_chart(chart, anchor)

    add_chart(ws, "Daily Pending by Bank Connector Code", bcc_rows, bcc_start, "E2")
    add_chart(ws, "Daily Pending by ReturnCode (BC)",     rc_rows,  rc_start,  f"E{rc_start+1}")


def build_findings_tab(ws, data, dates):
    """
    Data-driven findings: top BCC×date and RC×date entries across ALL dates,
    each with drill-down into ReturnCode and Sub Trans Type.
    """
    if not dates:
        ws["B1"] = "No data available."
        return

    # ── Column layout: A(#) | B(Finding) | C(ReturnCode/BCC drill) | D(SubTransType drill) ──
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 54
    ws.column_dimensions['D'].width = 54

    # ── Helpers ───────────────────────────────────────────────────────────────
    def fmt_top(series, total, n=5):
        top = series.sort_values(ascending=False).head(n)
        return " | ".join(
            f"{k}: {int(v):,} ({v/total*100:.0f}%)"
            for k, v in top.items() if v > 0
        )

    def spike_str(count, avg):
        if avg <= 0:
            return " (new)"
        pct = (count / avg - 1) * 100
        return f" ({pct:+.0f}% vs avg)" if abs(pct) >= 5 else ""

    latest = dates[-1]
    d_latest = data[data['_date'] == latest]
    d_hp     = data[~data['ReturnCode (BC)'].isin(LOW_PRIORITY_RC)]
    d_hp_lat = d_hp[d_hp['_date'] == latest]

    # ── Averages across all dates (for spike comparison) ─────────────────────
    bcc_date_df = data.groupby(['Bank Connector Code', '_date']).size().reset_index(name='n')
    bcc_avg = bcc_date_df.groupby('Bank Connector Code')['n'].mean().to_dict()

    rc_date_df = d_hp.groupby(['ReturnCode (BC)', '_date']).size().reset_index(name='n')
    rc_avg = rc_date_df.groupby('ReturnCode (BC)')['n'].mean().to_dict()

    # ── Build BCC finding items — top 5 BCCs on latest date ──────────────────
    bcc_counts = d_latest.groupby('Bank Connector Code').size().sort_values(ascending=False).head(5)
    bcc_findings = []
    for bcc, count in bcc_counts.items():
        avg  = bcc_avg.get(bcc, 0)
        d_sl = d_latest[d_latest['Bank Connector Code'] == bcc]

        # Col C: ReturnCode breakdown (aggregate)
        c_val = fmt_top(d_sl.groupby('ReturnCode (BC)').size(), count)

        # Col D: per-Sub Trans Type → ReturnCode breakdown (top 3 sub types × top 3 RCs)
        st_series = d_sl.groupby('Sub Trans Type').size().sort_values(ascending=False).head(3)
        d_lines = []
        for st_name, st_n in st_series.items():
            d_st = d_sl[d_sl['Sub Trans Type'] == st_name]
            rc_parts = " | ".join(
                f"{rc}: {int(n):,}"
                for rc, n in d_st.groupby('ReturnCode (BC)').size()
                              .sort_values(ascending=False).head(3).items()
            )
            d_lines.append(f"{st_name} ({int(st_n):,}) → {rc_parts}")
        d_val = "\n".join(d_lines)

        bcc_findings.append({
            'label': f"BCC: {bcc}  |  {latest}  |  {count:,} txns{spike_str(count, avg)}",
            'c_val': c_val,
            'd_val': d_val,
            'trend': 'down' if count >= avg * 1.2 else 'warn',
        })

    # ── Build RC finding items — top 5 high-priority RCs on latest date ──────
    rc_counts = d_hp_lat.groupby('ReturnCode (BC)').size().sort_values(ascending=False).head(5)
    rc_findings = []
    for rc, count in rc_counts.items():
        avg  = rc_avg.get(rc, 0)
        d_sl = d_hp_lat[d_hp_lat['ReturnCode (BC)'] == rc]

        # Col C: BCC breakdown (aggregate)
        bcc_series = d_sl.groupby('Bank Connector Code').size().sort_values(ascending=False)
        c_val = fmt_top(bcc_series, count)

        # Col D: per-BCC Sub Trans Type breakdown (top 3 BCCs × top 3 sub types)
        d_lines = []
        for bcc_name, bcc_n in bcc_series.head(3).items():
            d_bcc = d_sl[d_sl['Bank Connector Code'] == bcc_name]
            st_series = d_bcc.groupby('Sub Trans Type').size().sort_values(ascending=False).head(3)
            st_parts = " | ".join(f"{st}: {int(n):,}" for st, n in st_series.items())
            d_lines.append(f"{bcc_name} ({int(bcc_n):,}) → {st_parts}")
        d_val = "\n".join(d_lines)

        rc_findings.append({
            'label': f"RC: {rc}  |  {latest}  |  {count:,} txns{spike_str(count, avg)}",
            'c_val': c_val,
            'd_val': d_val,
            'trend': 'down' if count >= avg * 1.2 else 'warn',
        })

    # ── Write header ──────────────────────────────────────────────────────────
    ws.merge_cells("B1:D1")
    cl = ws["B1"]
    cl.value = f"Findings — Daily Pending Transactions  |  {dates[0]} → {dates[-1]}"
    cl.font = TITLE_FONT
    cl.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 32

    row = 2
    for col, hdr in [(2, "Finding  (BCC or ReturnCode | Date | Count | vs Avg)"),
                     (3, "Drill-down: ReturnCode (BC) / Bank Connector Code"),
                     (4, "BCC findings: Sub Trans Type → RC  |  RC findings: BCC → Sub Trans Type")]:
        cl = ws.cell(row=row, column=col, value=hdr)
        cl.font = H_FONT; cl.fill = HEADER_FILL
        cl.alignment = Alignment(horizontal='left', vertical='center')
        cl.border = B2
    ws.row_dimensions[row].height = 18
    row += 1

    # ── Section writer ────────────────────────────────────────────────────────
    def write_section(title, items, row):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cl = ws.cell(row=row, column=2, value=title)
        cl.font = SEC_FONT; cl.fill = HEADER_FILL
        cl.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 22
        row += 1

        for i, item in enumerate(items):
            f_b = DOWN_FILL  if item['trend'] == 'down' else WARN_FILL
            fnt = DOWN_FONT  if item['trend'] == 'down' else WARN_FONT
            f_c = ALT_FILL   if i % 2 == 0 else WHITE_FILL

            cl = ws.cell(row=row, column=2, value=item['label'])
            cl.font = fnt; cl.fill = f_b
            cl.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cl.border = B2

            cl = ws.cell(row=row, column=3, value=item['c_val'])
            cl.font = BODY_FONT; cl.fill = f_c
            cl.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cl.border = B2

            cl = ws.cell(row=row, column=4, value=item['d_val'])
            cl.font = BODY_FONT; cl.fill = f_c
            cl.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cl.border = B2

            ws.row_dimensions[row].height = 70
            row += 1

        return row + 1  # gap

    row = write_section(
        "Top 5 Bank Connector Codes  (ranked by volume × spike ratio, all dates)",
        bcc_findings, row)
    row = write_section(
        "Top 5 ReturnCodes (BC)  [low-priority 9202 / 5015 / 5007 / 5001 excluded]",
        rc_findings, row)

    # ── Legend ────────────────────────────────────────────────────────────────
    ws.cell(row=row, column=2).value = "Legend"
    ws.cell(row=row, column=2).font = Font(name="Arial", bold=True, size=10, color="1F3864")
    ws.row_dimensions[row].height = 18
    row += 1
    for lbl, fill, font in [
        ("🔴  Above average — investigate immediately", DOWN_FILL, DOWN_FONT),
        ("🟡  At or near average — monitor",            WARN_FILL, WARN_FONT),
    ]:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cl = ws.cell(row=row, column=2, value=lbl)
        cl.font = font; cl.fill = fill
        cl.alignment = Alignment(horizontal='left', vertical='center')
        cl.border = B2
        ws.row_dimensions[row].height = 18
        row += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def load_data(workspace_dir):
    """Load all CSVs from input/ folder, tag with date labels."""
    input_dir  = os.path.join(workspace_dir, "input")
    date_map_path = os.path.join(workspace_dir, "date_map.json")

    date_map = {}
    if os.path.exists(date_map_path):
        with open(date_map_path) as f:
            date_map = json.load(f)

    csv_files = sorted(glob.glob(os.path.join(input_dir, "translog_*.csv")))
    if not csv_files:
        raise FileNotFoundError(f"No translog_*.csv files found in {input_dir}")

    frames = []
    for fpath in csv_files:
        fname = os.path.basename(fpath)
        df = pd.read_csv(fpath)
        df['ReturnCode (BC)']     = df['ReturnCode (BC)'].str.strip().fillna('UNKNOWN')
        df['Sub Trans Type']      = df['Sub Trans Type'].str.strip().fillna('UNKNOWN')
        df['Bank Connector Code'] = df['Bank Connector Code'].str.strip().fillna('UNKNOWN')

        if fname in date_map:
            label = date_map[fname]
        else:
            # Derive from Trans Time column (dd/mm/yyyy → YYMMDD)
            sample = df['Trans Time'].dropna().iloc[0]
            parts = sample.split(' ')[0].split('/')
            label = parts[2][2:] + parts[1] + parts[0]  # YYMMDD

        df['_date'] = label
        frames.append(df)
        print(f"  Loaded {fname} → {label} ({len(df)} rows)")

    data = pd.concat(frames, ignore_index=True)
    data = data[~data['Sub Trans Type'].isin(EXCLUDE_SUB)]
    dates = sorted(set(data['_date'].unique()), key=lambda x: int(x))
    return data, dates


def main(workspace_dir, output_filename="Daily_Summary.xlsx"):
    print(f"Loading data from {workspace_dir}/input/...")
    data, dates = load_data(workspace_dir)
    print(f"  Dates found: {dates} | Total rows: {len(data)}")

    sort_date = dates[0]  # sort descending by earliest date
    output_path = os.path.join(workspace_dir, "output", output_filename)

    wb = Workbook()

    ws_findings = wb.active; ws_findings.title = "Findings"
    build_findings_tab(ws_findings, data, dates)

    ws_charts = wb.create_sheet("Charts")
    build_charts_tab(ws_charts, data, dates, sort_date)

    ws_bcc = wb.create_sheet("By Bank Connector Code")
    build_bcc_tab(ws_bcc, data, dates, sort_date)

    ws_rc = wb.create_sheet("By ReturnCode (BC)")
    build_rc_tab(ws_rc, data, dates, sort_date)

    ws_detail = wb.create_sheet("By Sub Trans Type (Detail)")
    build_detail_tab(ws_detail, data, dates, sort_date)

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_daily_summary.py <workspace_dir> [output_filename]")
        sys.exit(1)
    workspace_dir   = sys.argv[1]
    output_filename = sys.argv[2] if len(sys.argv) > 2 else "Daily_Summary.xlsx"
    main(workspace_dir, output_filename)
