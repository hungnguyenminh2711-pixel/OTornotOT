"""
Pending Transactions Summary Generator
Reads a translog CSV and produces a 4-tab Excel summary.
Usage: python generate_summary.py <input_csv> <output_xlsx>
"""
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
EXCLUDE_SUB = {
    'LINK', 'QUERY_BALANCE_USER', 'REGISTER_QUERY_BALANCE', 'UNLINK', 'VERIFY_OTP',
    'UNREGISTER_QUERY_BALANCE'
}
DETAIL_SUBS = [
    'DOMESTIC_TRANSFER_FUND_2_ACCOUNT', 'QR_PUSH_PAYMENT', 'WITHDRAW_BY_TOKEN'
]

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
SECTION_FILL = PatternFill("solid", start_color="2E75B6")
TOTAL_FILL   = PatternFill("solid", start_color="BDD7EE")
ALT_FILL     = PatternFill("solid", start_color="D9E1F2")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")

H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
S_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
N_FONT = Font(name="Arial", size=9)
B_FONT = Font(name="Arial", bold=True, size=9)
T_FONT = Font(name="Arial", bold=True, size=9, color="FFFFFF")

thin  = Side(style='thin',   color="AAAAAA")
med   = Side(style='medium', color="2E75B6")
T_BDR = Border(left=thin, right=thin, top=thin, bottom=thin)
M_BDR = Border(left=med,  right=med,  top=med,  bottom=med)

C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left',   vertical='center')
R = Alignment(horizontal='right',  vertical='center')
LW = Alignment(horizontal='left',  vertical='top', wrap_text=True)


def hc(ws, r, c, v):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = H_FONT; cl.fill = HEADER_FILL; cl.alignment = C; cl.border = T_BDR


def tc(ws, r, c, v):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = T_FONT; cl.fill = HEADER_FILL; cl.alignment = R; cl.border = T_BDR


def sc(ws, r, c, v, ncols):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + ncols - 1)
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = S_FONT; cl.fill = SECTION_FILL; cl.alignment = L; cl.border = M_BDR


def write_table(ws, start_row, start_col, row_vals, col_vals, lookup, row_label):
    """
    Grand Total is 2nd column. Layout:
      col A = row_label | col B = Grand Total | col C.. = data cols
    Returns next free row (with 2-row gap).
    """
    gt_col   = start_col + 1
    data_col = start_col + 2

    hc(ws, start_row, start_col, row_label)
    hc(ws, start_row, gt_col, "Grand Total")
    for j, cv in enumerate(col_vals):
        hc(ws, start_row, data_col + j, cv)

    for i, rv in enumerate(row_vals):
        r = start_row + 1 + i
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        cl = ws.cell(row=r, column=start_col, value=rv)
        cl.font = B_FONT; cl.fill = fill; cl.alignment = L; cl.border = T_BDR

        refs = []
        for j, cv in enumerate(col_vals):
            c = data_col + j
            val = lookup.get((rv, cv), 0)
            cl = ws.cell(row=r, column=c, value=val)
            cl.font = N_FONT; cl.fill = fill; cl.alignment = R; cl.border = T_BDR
            refs.append(get_column_letter(c))

        cl = ws.cell(row=r, column=gt_col,
                     value="=" + "+".join(f"{x}{r}" for x in refs))
        cl.font = B_FONT; cl.fill = TOTAL_FILL; cl.alignment = R; cl.border = T_BDR

    tr = start_row + 1 + len(row_vals)
    cl = ws.cell(row=tr, column=start_col, value="Grand Total")
    cl.font = T_FONT; cl.fill = HEADER_FILL; cl.alignment = L; cl.border = T_BDR
    gl = get_column_letter(gt_col)
    tc(ws, tr, gt_col, f"=SUM({gl}{start_row+1}:{gl}{tr-1})")
    for j in range(len(col_vals)):
        c = data_col + j
        l = get_column_letter(c)
        tc(ws, tr, c, f"=SUM({l}{start_row+1}:{l}{tr-1})")

    # column widths
    ws.column_dimensions[get_column_letter(start_col)].width = max(
        ws.column_dimensions[get_column_letter(start_col)].width or 0, 26)
    ws.column_dimensions[get_column_letter(gt_col)].width = max(
        ws.column_dimensions[get_column_letter(gt_col)].width or 0, 14)
    for j, cv in enumerate(col_vals):
        c = data_col + j
        ws.column_dimensions[get_column_letter(c)].width = max(
            ws.column_dimensions[get_column_letter(c)].width or 0,
            min(len(str(cv)) + 2, 38))

    return tr + 2


# ── Tab builders ──────────────────────────────────────────────────────────────

def build_bcc_pivot(ws, df_in, col_field, col_label, title, exclude=None):
    """BCC (rows, sorted desc by total) × col_field, Grand Total 2nd col."""
    d = df_in.copy()
    if exclude:
        d = d[~d[col_field].isin(exclude)]
    cols   = sorted(d[col_field].unique())
    pivot  = d.groupby(['Bank Connector Code', col_field]).size().reset_index(name='n')
    totals = pivot.groupby('Bank Connector Code')['n'].sum().sort_values(ascending=False)
    connectors = totals.index.tolist()
    lookup = pivot.set_index(['Bank Connector Code', col_field])['n'].to_dict()

    gt_col   = 2
    data_col = 3

    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="1F3864")
    ws["A1"].alignment = L

    hc(ws, 2, 1, "Bank Connector Code")
    hc(ws, 2, gt_col, "Grand Total")
    for j, cv in enumerate(cols):
        hc(ws, 2, data_col + j, cv)

    for i, connector in enumerate(connectors):
        r = 3 + i
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        cl = ws.cell(row=r, column=1, value=connector)
        cl.font = B_FONT; cl.fill = fill; cl.alignment = L; cl.border = T_BDR

        refs = []
        for j, cv in enumerate(cols):
            c = data_col + j
            val = lookup.get((connector, cv), 0)
            cl = ws.cell(row=r, column=c, value=val)
            cl.font = N_FONT; cl.fill = fill; cl.alignment = R; cl.border = T_BDR
            refs.append(get_column_letter(c))

        cl = ws.cell(row=r, column=gt_col,
                     value="=" + "+".join(f"{x}{r}" for x in refs))
        cl.font = B_FONT; cl.fill = TOTAL_FILL; cl.alignment = R; cl.border = T_BDR

    tr = 3 + len(connectors)
    cl = ws.cell(row=tr, column=1, value="Grand Total")
    cl.font = T_FONT; cl.fill = HEADER_FILL; cl.alignment = L; cl.border = T_BDR
    gl = get_column_letter(gt_col)
    tc(ws, tr, gt_col, f"=SUM({gl}3:{gl}{tr-1})")
    for j in range(len(cols)):
        c = data_col + j; l = get_column_letter(c)
        tc(ws, tr, c, f"=SUM({l}3:{l}{tr-1})")

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions[get_column_letter(gt_col)].width = 14
    for j, cv in enumerate(cols):
        ws.column_dimensions[get_column_letter(data_col + j)].width = max(
            18, min(len(str(cv)) + 2, 35))
    ws.freeze_panes = "C3"


def build_detail_tab(ws, d):
    """Overview + per-sub-trans-type BCC breakdowns."""
    return_codes = sorted(d['ReturnCode (BC)'].unique())
    sub_types    = sorted(d['Sub Trans Type'].unique())

    ws["A1"] = "Overview: ReturnCode (BC) × Sub Trans Type"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = L

    lookup_main = d.groupby(['ReturnCode (BC)', 'Sub Trans Type']).size().to_dict()
    current_row = write_table(ws, 2, 1, return_codes, sub_types, lookup_main, "ReturnCode (BC)")

    for st in sub_types:
        d_st     = d[d['Sub Trans Type'] == st]
        bcc_list = sorted(d_st['Bank Connector Code'].unique())
        rc_list  = sorted(d_st['ReturnCode (BC)'].unique())

        ncols = max(len(bcc_list), len(rc_list)) + 3
        sc(ws, current_row, 1, f"▸  {st}", ncols)
        current_row += 1

        ws.cell(row=current_row, column=1).value = f"BCC – {st}"
        ws.cell(row=current_row, column=1).font = Font(
            name="Arial", bold=True, size=10, color="2E75B6")
        ws.cell(row=current_row, column=1).alignment = L
        current_row += 1

        lk_bcc_st = d_st.groupby(['Bank Connector Code', 'Sub Trans Type']).size().to_dict()
        current_row = write_table(ws, current_row, 1, bcc_list, [st], lk_bcc_st, "Bank Connector Code")

        ws.cell(row=current_row, column=1).value = f"BCC – ReturnCode (BC) for {st}"
        ws.cell(row=current_row, column=1).font = Font(
            name="Arial", bold=True, size=10, color="2E75B6")
        ws.cell(row=current_row, column=1).alignment = L
        current_row += 1

        lk_bcc_rc = d_st.groupby(['Bank Connector Code', 'ReturnCode (BC)']).size().to_dict()
        current_row = write_table(ws, current_row, 1, bcc_list, rc_list, lk_bcc_rc, "Bank Connector Code")

    ws.freeze_panes = "C3"


def build_findings_tab(ws, df):
    """Auto-generate key findings for each table."""
    TITLE_FONT   = Font(name="Arial", bold=True, size=14, color="1F3864")
    SEC_FONT     = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    LABEL_FONT   = Font(name="Arial", bold=True, size=10, color="2E75B6")
    BODY_FONT    = Font(name="Arial", size=10)
    SEC_FILL     = PatternFill("solid", start_color="1F3864")
    thin2        = Side(style='thin', color="CCCCCC")
    B2           = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)
    LC           = Alignment(horizontal='left', vertical='center')

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 80

    d_sub = df[~df['Sub Trans Type'].isin(EXCLUDE_SUB)]
    d_det = df[df['Sub Trans Type'].isin(DETAIL_SUBS)]

    # --- derive findings dynamically ---
    sub_totals = d_sub.groupby('Sub Trans Type').size().sort_values(ascending=False)
    top1_st, top1_n = sub_totals.index[0], sub_totals.iloc[0]
    top2_st, top2_n = sub_totals.index[1], sub_totals.iloc[1]
    total_sub = sub_totals.sum()
    pct1 = top1_n / total_sub * 100
    pct2 = top2_n / total_sub * 100

    bcc_total = d_sub.groupby('Bank Connector Code').size()
    bcc_st    = d_sub.groupby(['Bank Connector Code', 'Sub Trans Type']).size()
    single_type = [b for b in bcc_total.index
                   if bcc_st[b].max() / bcc_total[b] == 1.0]

    rc_totals = df.groupby('ReturnCode (BC)').size().sort_values(ascending=False)
    rc1, rc1n = rc_totals.index[0], rc_totals.iloc[0]
    rc2, rc2n = rc_totals.index[1], rc_totals.iloc[1]
    total_rc  = rc_totals.sum()
    rc1_pct   = rc1n / total_rc * 100
    rc2_pct   = rc2n / total_rc * 100
    top3_bcc  = df.groupby('Bank Connector Code').size().sort_values(ascending=False).head(3)
    top3_n    = top3_bcc.sum()
    top3_pct  = top3_n / total_rc * 100

    dt_rc = d_det.groupby(['Sub Trans Type', 'ReturnCode (BC)']).size()
    dt_bcc = d_det.groupby(['Sub Trans Type', 'Bank Connector Code']).size()

    findings = [
        ("Tab: By Sub Trans Type", [
            (f"{top1_st} dominates volume",
             f"{top1_st} accounts for {top1_n:,} transactions ({pct1:.0f}% of filtered total). "
             f"Most BCCs handling this type do so exclusively."),
            (f"{top2_st} is the 2nd largest type",
             f"{top2_st} contributes {top2_n:,} transactions ({pct2:.0f}%). "
             f"Together the top 2 types account for {pct1+pct2:.0f}% of all pending transactions."),
            (f"{len(single_type)} BCCs are single-type connectors",
             f"{', '.join(single_type[:5])}{'...' if len(single_type)>5 else ''} each handle "
             f"only one Sub Trans Type (100% concentration), indicating highly specialised connectors."),
            ("Low-volume types warrant monitoring",
             f"{', '.join(sub_totals[sub_totals==1].index.tolist())} each have only 1 transaction. "
             f"These may be newly onboarded flows or isolated incidents."),
        ]),
        ("Tab: By ReturnCode (BC)", [
            (f"{rc1.split(' ')[0]} is the #1 error",
             f"{rc1n:,} transactions ({rc1_pct:.1f}%) failed with {rc1}. "
             f"The bank never sent a callback — investigate IPN endpoint reachability."),
            (f"{rc2.split('(')[0].strip()} is 2nd largest",
             f"{rc2n:,} transactions ({rc2_pct:.1f}%) hit {rc2}. "
             f"Suggests persistent connectivity issues on specific bank connectors."),
            (f"Top 3 BCCs drive {top3_pct:.0f}% of all errors",
             f"{', '.join(top3_bcc.index.tolist())} together generate {top3_n:,} transactions "
             f"— the highest-priority targets for investigation."),
            ("Rare error codes signal config bugs",
             f"Low-frequency codes like BANK_PARAMETER_INVALID and BANK_DATA_CONVERTION_EXCEPTION "
             f"appear in single digits. These likely indicate data mapping or integration issues."),
        ]),
        ("Tab: By Sub Trans Type (Detail)", []),
    ]

    # Dynamic detail findings
    detail_findings = []
    for st in DETAIL_SUBS:
        if st not in dt_rc.index.get_level_values(0):
            continue
        rc_dist  = dt_rc[st].sort_values(ascending=False)
        bcc_dist = dt_bcc[st].sort_values(ascending=False)
        top_rc   = rc_dist.index[0]
        top_rc_n = rc_dist.iloc[0]
        top_bcc  = bcc_dist.index[0]
        top_bcc_n = bcc_dist.iloc[0]
        total_st = rc_dist.sum()
        pct_rc   = top_rc_n / total_st * 100
        pct_bcc  = top_bcc_n / total_st * 100
        detail_findings.append((
            f"{st}: {top_rc.split('(')[0].strip()} dominates",
            f"{top_rc_n} of {total_st} transactions ({pct_rc:.0f}%) failed with {top_rc}. "
            f"{top_bcc} drives {top_bcc_n} transactions ({pct_bcc:.0f}%) for this type."
        ))
    findings[2][1].extend(detail_findings[:5])

    ws["B1"] = f"Key Findings — Pending Transactions"
    ws["B1"].font = TITLE_FONT; ws["B1"].alignment = LC
    ws.row_dimensions[1].height = 28

    row = 3
    for section_title, items in findings:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        cl = ws.cell(row=row, column=2, value=section_title)
        cl.font = SEC_FONT; cl.fill = SEC_FILL; cl.alignment = LC
        ws.row_dimensions[row].height = 20
        row += 1

        for i, (label, body) in enumerate(items):
            fill2 = ALT_FILL if i % 2 == 0 else WHITE_FILL
            ws.cell(row=row, column=1, value=f"{i+1}").font = Font(
                name="Arial", bold=True, size=10, color="888888")
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')

            cl = ws.cell(row=row, column=2, value=label)
            cl.font = LABEL_FONT; cl.fill = fill2; cl.alignment = LC; cl.border = B2

            cl = ws.cell(row=row, column=3, value=body)
            cl.font = BODY_FONT; cl.fill = fill2; cl.alignment = LW; cl.border = B2

            ws.row_dimensions[row].height = 42
            row += 1
        row += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def main(input_csv, output_xlsx):
    df = pd.read_csv(input_csv)
    df['ReturnCode (BC)']    = df['ReturnCode (BC)'].fillna('UNKNOWN').astype(str).str.strip()
    df['Sub Trans Type']     = df['Sub Trans Type'].fillna('UNKNOWN').astype(str).str.strip()
    df['Bank Connector Code'] = df['Bank Connector Code'].fillna('UNKNOWN').astype(str).str.strip()

    # Apply exclusion filter for all tabs
    df_filtered = df[~df['Sub Trans Type'].isin(EXCLUDE_SUB)]
    d_det = df_filtered[df_filtered['Sub Trans Type'].isin(DETAIL_SUBS)]

    wb = Workbook()

    # Tab order: Findings | By Sub Trans Type | By ReturnCode (BC) | Detail
    ws_findings = wb.active
    ws_findings.title = "Findings"
    build_findings_tab(ws_findings, df_filtered)

    ws1 = wb.create_sheet("By Sub Trans Type")
    build_bcc_pivot(ws1, df_filtered, 'Sub Trans Type', 'Sub Trans Type',
                    "Bank Connector Code × Sub Trans Type")

    ws2 = wb.create_sheet("By ReturnCode (BC)")
    build_bcc_pivot(ws2, df_filtered, 'ReturnCode (BC)', 'ReturnCode (BC)',
                    "Bank Connector Code × ReturnCode (BC)")

    ws3 = wb.create_sheet("By Sub Trans Type (Detail)")
    build_detail_tab(ws3, d_det)

    wb.save(output_xlsx)
    print(f"Saved: {output_xlsx}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_summary.py <input.csv> <output.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
